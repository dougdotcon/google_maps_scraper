import time
import re
import pandas as pd
import threading
import queue
import csv
import os
from typing import List, Dict, Set
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from browser import BrowserManager
import tkinter as tk
from tkinter import ttk, messagebox
import sys

# Thread-safe queues
url_queue = queue.Queue()
result_queue = queue.Queue()

class GoogleMapsScraper:
    def __init__(self, headless=False):
        self.driver = BrowserManager.create_driver(headless)
        self.wait = WebDriverWait(self.driver, 20)
        
    def close(self):
        if self.driver:
            self.driver.quit()

    def clean_phone(self, phone):
        if not phone: return ""
        return re.sub(r'[^\d\(\)\-\+\s]', '', phone).strip()

    def search_location(self, address):
        print(f"Searching for location: {address}")
        self.driver.get('https://www.google.com/maps/')
        try:
            search_box = self.wait.until(
                EC.presence_of_element_located((By.XPATH, '//input[@id="searchboxinput"]'))
            )
            search_box.clear()
            search_box.send_keys(address)
            search_box.send_keys(Keys.ENTER)
            time.sleep(5) # Wait for zoom
            return True
        except Exception as e:
            print(f"Error searching location: {e}")
            return False

    def search_nearby(self, keyword="empresas"):
        print(f"Searching nearby for: {keyword}")
        try:
            search_box = self.wait.until(
                EC.presence_of_element_located((By.XPATH, '//input[@id="searchboxinput"]'))
            )
            search_box.click()
            search_box.send_keys(Keys.CONTROL + "a")
            search_box.send_keys(Keys.DELETE)
            search_box.send_keys(keyword)
            search_box.send_keys(Keys.ENTER)
            time.sleep(5)
        except Exception as e:
            print(f"Error initiating nearby search: {e}")

    def collect_links(self, max_leads=50) -> List[str]:
        """
        Scrolls the feed and collects direct links to the places.
        Does NOT visit them.
        """
        links = set()
        print(f"Collecting up to {max_leads} links...")
        
        xpath_result = '//a[contains(@href, "/place/")]'
        consecutive_failures = 0
        
        while len(links) < max_leads:
            try:
                elements = self.driver.find_elements(By.XPATH, xpath_result)
                if not elements:
                    print("No elements found during collection.")
                    break
                
                new_found = False
                for element in elements:
                    try:
                        href = element.get_attribute("href")
                        if href and href not in links:
                             links.add(href)
                             new_found = True
                             if len(links) >= max_leads: 
                                 break
                    except:
                        continue
                
                if len(links) >= max_leads:
                    break

                # Scroll
                if not new_found:
                     print("No new links visible, scrolling...")
                
                try:
                    feed = self.driver.find_element(By.XPATH, '//div[@role="feed"]')
                    self.driver.execute_script("arguments[0].scrollTop += 1000;", feed)
                except: 
                     self.driver.execute_script("window.scrollBy(0, 1000);")
                
                time.sleep(2)
                
                # Check end of list detection (naive)
                if not new_found:
                    consecutive_failures += 1
                else:
                    consecutive_failures = 0
                    
                if consecutive_failures > 5:
                    print("Collection seems to have reached the end.")
                    break

            except Exception as e:
                print(f"Collection error: {e}")
                break
                
        print(f"Collected {len(links)} unique links.")
        return list(links)

    def scrape_url(self, url):
        """
        Navigates directly to the URL and extracts details.
        """
        extracted = {
            'name': '',
            'address': '',
            'phone': '',
            'website': '',
            'link': url,
            'status': 'success'
        }
        
        try:
            self.driver.get(url)
            # Wait for H1 or some indicator
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, 'h1'))
                )
            except:
                pass # Try anyway
                
            # Extract basic info similar to before
            try:
                h1s = self.driver.find_elements(By.TAG_NAME, 'h1')
                for h1 in h1s:
                    t = h1.text
                    if t and len(t) > 2 and t not in ["Resultados", "Principal", "Patrocinado"]:
                        extracted['name'] = t
                        break
            except: pass
            
            try:
                addr_btn = self.driver.find_element(By.XPATH, '//button[contains(@data-item-id, "address")]')
                extracted['address'] = addr_btn.get_attribute("aria-label").replace("Endereço: ", "")
            except: pass
            
            try:
                phone_btn = self.driver.find_element(By.XPATH, '//button[contains(@data-item-id, "phone")]')
                extracted['phone'] = self.clean_phone(phone_btn.get_attribute("aria-label").replace("Telefone: ", ""))
            except: pass
            
            try:
                web_btn = self.driver.find_element(By.XPATH, '//a[contains(@data-item-id, "authority")]')
                extracted['website'] = web_btn.get_attribute('href')
            except: pass
            
            return extracted
            
        except Exception as e:
            print(f"Error scraping URL {url}: {e}")
            extracted['status'] = 'error'
            return extracted

def worker_scraper(headless):
    # Each worker has its own driver instance
    scraper = None
    try:
        scraper = GoogleMapsScraper(headless=headless)
        while True:
            try:
                task = url_queue.get(timeout=3) # Wait 3 sec for new task
            except queue.Empty:
                break # Queue is empty and we are done
            
            keyword, url = task
            try:
                data = scraper.scrape_url(url)
                data['keyword'] = keyword
                
                # Check if valid data
                if data['name']:
                    print(f"Extracted: {data['name']}")
                    result_queue.put(data)
                else:
                    print(f"Failed to extract name for {url}")
            except Exception as e:
                print(f"Worker error on {url}: {e}")
            finally:
                url_queue.task_done()
    except Exception as e:
        print(f"Worker thread init failed: {e}")
    finally:
        if scraper:
            scraper.close()

def worker_saver(filename, stop_event):
    # Check if file exists to write header
    file_exists = os.path.isfile(filename)
    
    with open(filename, mode='a', newline='', encoding='utf-8-sig') as f:
        fieldnames = ['keyword', 'name', 'address', 'phone', 'website', 'link', 'status']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        
        if not file_exists:
            writer.writeheader()
            
        while not stop_event.is_set() or not result_queue.empty():
            try:
                data = result_queue.get(timeout=0.5)
                writer.writerow(data)
                f.flush() # Ensure it's written immediately
                result_queue.task_done()
            except queue.Empty:
                continue
            except Exception as e:
                print(f"Saver error: {e}")

def get_user_config():
    config = {}
    
    def on_submit():
        config['address'] = address_entry.get().strip()
        config['keywords'] = [k.strip() for k in keywords_entry.get().split(',') if k.strip()]
        try:
            config['max_leads'] = int(max_leads_entry.get().strip())
        except ValueError:
            messagebox.showerror("Erro", "Max Leads deve ser inteiro.")
            return
        
        try:
            config['workers'] = int(workers_entry.get().strip())
            if config['workers'] < 1: raise ValueError
        except ValueError:
            messagebox.showerror("Erro", "Workers deve ser inteiro > 0.")
            return

        config['headless'] = headless_var.get()
        
        if not config['address'] or not config['keywords']:
            messagebox.showerror("Erro", "Campos obrigatórios vazios.")
            return
            
        root.destroy()

    def on_cancel():
        config.clear()
        root.destroy()
        sys.exit()

    root = tk.Tk()
    root.title("Configuração Scraper (Paralelo)")
    root.geometry("500x450")
    
    frame = ttk.Frame(root, padding="20")
    frame.pack(fill=tk.BOTH, expand=True)

    ttk.Label(frame, text="Endereço Central:").pack(anchor=tk.W)
    address_entry = ttk.Entry(frame, width=50)
    address_entry.insert(0, "Avenida Embaixador Abelardo Bueno, 3401 – Rio de Janeiro")
    address_entry.pack(fill=tk.X, pady=5)

    ttk.Label(frame, text="Termos (ex: bar, padaria):").pack(anchor=tk.W)
    keywords_entry = ttk.Entry(frame, width=50)
    keywords_entry.insert(0, "restaurante, mercado")
    keywords_entry.pack(fill=tk.X, pady=5)

    ttk.Label(frame, text="Max Leads por termo:").pack(anchor=tk.W)
    max_leads_entry = ttk.Entry(frame, width=10)
    max_leads_entry.insert(0, "50")
    max_leads_entry.pack(anchor=tk.W, pady=5)
    
    ttk.Label(frame, text="Número de Workers (Navegadores Simultâneos):").pack(anchor=tk.W)
    ttk.Label(frame, text="Cuidado: Mais workers consomem mais RAM/CPU", font=("Arial", 8), foreground="red").pack(anchor=tk.W)
    workers_entry = ttk.Entry(frame, width=10)
    workers_entry.insert(0, "2")
    workers_entry.pack(anchor=tk.W, pady=5)

    headless_var = tk.BooleanVar(value=True) # Default True for parallel
    ttk.Checkbutton(frame, text="Headless (Recomendado para Paralelo)", variable=headless_var).pack(anchor=tk.W, pady=10)

    ttk.Button(frame, text="Iniciar", command=on_submit).pack(pady=10)

    root.mainloop()
    return config

if __name__ == "__main__":
    config = get_user_config()
    if not config: sys.exit()
    
    print(f"=== Configuração: {config} ===")
    
    # Generate Filename (CSV first)
    timestamp = int(time.time())
    csv_filename = f"leads_temp_{timestamp}.csv"
    xlsx_filename = f"leads_final_{timestamp}.xlsx"
    
    # 1. Start Collector (Main Thread or Single Worker)
    print(">>> Iniciando Coletor de Links...")
    collector = GoogleMapsScraper(headless=config['headless'])
    
    total_links = 0
    try:
        if collector.search_location(config['address']):
            for keyword in config['keywords']:
                print(f"--- Coletando links para: {keyword} ---")
                collector.search_nearby(keyword)
                links = collector.collect_links(max_leads=config['max_leads'])
                for link in links:
                    url_queue.put((keyword, link))
                total_links += len(links)
    except Exception as e:
        print(f"Erro na coleta: {e}")
    finally:
        collector.close()
        
    print(f">>> Coleta finalizada. Total de links: {total_links}")
    if total_links == 0:
        messagebox.showinfo("Fim", "Nenhum link encontrado.")
        sys.exit()

    # 2. Start Saver Thread
    stop_saver = threading.Event()
    saver_thread = threading.Thread(target=worker_saver, args=(csv_filename, stop_saver))
    saver_thread.start()
    
    # 3. Start Worker Threads (Consumers)
    print(f">>> Iniciando {config['workers']} workers de extração...")
    threads = []
    for i in range(config['workers']):
        t = threading.Thread(target=worker_scraper, args=(config['headless'],))
        t.start()
        threads.append(t)
        
    # 4. Wait for completion
    for t in threads:
        t.join()
        
    print(">>> Todos os workers finalizaram.")
    
    # Stop Saver
    stop_saver.set()
    saver_thread.join()
    
    # 5. Convert CSV to Excel
    print(">>> Convertendo para Excel...")
    try:
        if os.path.exists(csv_filename):
            # Read CSV, forcing phone to be string to avoid float conversion issues
            df = pd.read_csv(csv_filename, dtype={'phone': str})
            
            # --- NOVA LÓGICA: STATUS SITE ---
            if 'website' in df.columns:
                df['STATUS SITE'] = df['website'].apply(
                    lambda x: '✅ COM SITE' if pd.notna(x) and str(x).strip() != '' else '❌ SEM SITE'
                )

            # --- NOVA LÓGICA: STATUS SEO ADS ---
            # Mede eficiência pela quantidade de separadores "|" no nome
            if 'name' in df.columns:
                def analyze_seo(name_val):
                    if pd.isna(name_val): return 0
                    count = str(name_val).count('|')
                    if count == 0:
                        return "Básico"
                    else:
                        return f"Otimizado ({count} tags)"
                
                df['STATUS SEO ADS'] = df['name'].apply(analyze_seo)
            
            # --- LÓGICA: WHATSAPP LINK ---
            if 'phone' in df.columns:
                def make_whatsapp_link(phone_raw):
                    if pd.isna(phone_raw) or not str(phone_raw).strip():
                        return ""
                    s_phone = str(phone_raw).strip()
                    digits = re.sub(r'\D', '', s_phone)
                    
                    if not digits:
                        return s_phone
                        
                    target_num = digits
                    if len(digits) in [10, 11] and not digits.startswith('55'):
                         target_num = '55' + digits
                    
                    link = f"https://wa.me/{target_num}"
                    safe_phone = s_phone.replace('"', '""')
                    return f'=HYPERLINK("{link}", "{safe_phone}")'

                df['phone'] = df['phone'].apply(make_whatsapp_link)

            # Reordenar colunas para melhor visualização
            desired_order = ['keyword', 'name', 'STATUS SEO ADS', 'phone', 'address', 'STATUS SITE', 'website', 'link']
            # Garante que pegamos todas as colunas que existem no DF, na ordem desejada + as sobras
            existing_cols = [c for c in desired_order if c in df.columns]
            remaining_cols = [c for c in df.columns if c not in existing_cols]
            df = df[existing_cols + remaining_cols]

            # Save with formatting (AutoFilter + Column Widths)
            try:
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Font, PatternFill, Alignment
                
                with pd.ExcelWriter(xlsx_filename, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Leads')
                    worksheet = writer.sheets['Leads']
                    
                    # Apply AutoFilter
                    worksheet.auto_filter.ref = worksheet.dimensions
                    
                    # Header Style
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")

                    # Adjust Column Widths
                    for i, column in enumerate(df.columns, 1):
                        col_letter = get_column_letter(i)
                        
                        # Calculate best width
                        max_len = 0
                        
                        # Check header length
                        if column:
                            max_len = len(str(column))
                        
                        # Check first 50 rows data length to save time if large
                        try:
                            # Convert series to string, handle None
                            series = df[column].astype(str)
                            if not series.empty:
                                data_max = series.map(len).max()
                                if pd.notna(data_max) and data_max > max_len:
                                    max_len = data_max
                        except: pass
                        
                        # Limits
                        adjusted_width = (max_len + 3) * 1.1
                        if adjusted_width > 60: adjusted_width = 60
                        if adjusted_width < 12: adjusted_width = 12
                        
                        worksheet.column_dimensions[col_letter].width = adjusted_width
                        
            except Exception as format_error:
                print(f"Aviso: Erro na formatação do Excel ({format_error}). Salvando padrão.")
                df.to_excel(xlsx_filename, index=False)

            print(f"Arquivo final salvo: {xlsx_filename}")
            try:
                os.remove(csv_filename) # Clean up temp csv
            except: pass
            
            messagebox.showinfo("Sucesso", f"Processo concluído!\nSalvo em: {xlsx_filename}\nTotal extraído: {len(df)}")
        else:
            print("Nenhum dado foi salvo no CSV.")
    except Exception as e:
        print(f"Erro ao converter Excel: {e}")
